#!/usr/bin/env python   
# -*- coding: utf-8 -*- 
from win32com.client import Dispatch
import win32com.client
import time
import os
import os.path
import openpyxl
import sys
import csv
import datetime
import zipfile
from zipfile import *
import xlrd
import xdrlib,sys
import xlwt
import urllib.request
import urllib.parse
import http.cookiejar
import re
import datetime
import json
from bs4 import BeautifulSoup
import ssl
from xlutils.copy import copy



def open_excel(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception(e):
        print(str(e))
           
def excel_get_zone_or_shipment(file,colnameindex,by_name,DaysRange):
    current_date = 0
    current_date = time.time() - 86400*int(DaysRange)
    timeStr = time.strftime("%y-%m-%d",time.localtime(current_date))
    data = open_excel(file)
    table = data.sheet_by_name(by_name)
    nrows = table.nrows
    colnames = table.row_values(colnameindex)
    list = []
    list2 = []
    for rownum in range(0,nrows):
        row = table.row_values(rownum)
        if row:
            if(str(table.cell(rownum,0)) == "text:'20"+str(timeStr)+"'"):
                app = {}
                for i in range(len(colnames)):
                    app[colnames[i]] = row[i]
                list.append(app)              
    
    for i in range(len(list)):
            for j in range(len(list[i])):
                    list2.append(list[i][colnames[j]])
    return list2

def excel_get_user(file,colnameindex = 0, by_name=u'Sheet0'):
    data = open_excel(file)
    table = data.sheet_by_name(by_name)
    nrows = table.nrows
    colnames = table.row_values(colnameindex)
    list = []
    list2 = []
    for rownum in range(0,nrows):
        row = table.row_values(rownum)
        if row:
            app = {}
            for i in range(len(colnames)):
                app[colnames[i]] = row[i]
            list.append(app)
    for i in range(len(list)):
        for j in range(len(list[i])):
            list2.append(list[i][colnames[j]])
    return list2[3:]

#全局变量
aczone_br_daily = 0
aczone_br_edi = 0
aczone_br_online = 0
aczone_si_daily = 0
aczone_si_edi = 0
aczone_si_online = 0

stdzone_br_daily = 0
stdzone_br_edi = 0
stdzone_br_online = 0
stdzone_si_daily = 0
stdzone_si_edi = 0
stdzone_si_online = 0

def calc(list_aczone,list_stdzone):
    
    list_daily = []
    list_calc = []
    
    #声明
    global aczone_br_daily
    global aczone_br_edi
    global aczone_br_online
    global aczone_si_daily
    global aczone_si_edi
    global aczone_si_online

    global stdzone_br_daily
    global stdzone_br_edi
    global stdzone_br_online
    global stdzone_si_daily
    global stdzone_si_edi
    global stdzone_si_online

    for i in range(len(list_aczone)):
        if list_aczone[i] == 'BR':
            aczone_br_daily = list_aczone[i-1]
            aczone_br_edi = list_aczone[i+1]
            aczone_br_online = list_aczone[i+2]
            
        if list_aczone[i] == 'SI':
            aczone_si_daily = list_aczone[i-1]
            aczone_si_edi = list_aczone[i+1]
            aczone_si_online = list_aczone[i+2]

    for j in range(len(list_stdzone)):
        if list_stdzone[j] == 'BR':
            stdzone_br_daily = list_stdzone[j-1]
            stdzone_br_edi = list_stdzone[j+1]
            stdzone_br_online = list_stdzone[j+2]
            
        if list_stdzone[j] == 'SI':
            stdzone_si_daily = list_stdzone[j-1]
            stdzone_si_edi = list_stdzone[j+1]
            stdzone_si_online = list_stdzone[j+2]
    
    list_daily.append(stdzone_br_daily)
    list_daily.append(aczone_br_daily)
    list_daily.append(stdzone_si_daily)
    list_daily.append(aczone_si_daily)

    list_calc.append(int(aczone_br_edi)+int(stdzone_br_edi))
    list_calc.append(int(aczone_br_online)+int(stdzone_br_online))
    list_calc.append(int(aczone_si_online)+int(stdzone_si_online))
    list_calc.append(int(aczone_si_edi)+int(stdzone_si_edi))
    list_calc.append(int(aczone_br_edi)+int(stdzone_br_edi)+int(aczone_br_online)+int(stdzone_br_online)+int(aczone_si_online)+int(stdzone_si_online)+int(aczone_si_edi)+int(stdzone_si_edi))

    result = [list_daily,list_calc]
    
    return result

############爬虫##########

def getOpener(head):
    #deal with the Cookies
    cj = http.cookiejar.CookieJar()
    pro = urllib.request.HTTPCookieProcessor(cj)
    opener = urllib.request.build_opener(pro)
    header = []
    for key,value in head.items():
        elem = (key,value)
        header.append(elem)
    opener.addheaders = header
    return opener

def getTimeSec(x):
    dt = datetime.datetime.now()
    #x为要减的天数
    begin_day = dt + datetime.timedelta(days = (0-int(x)))
    end_day = dt + datetime.timedelta(days=(0-int(x)+1))
    current_year = int('20'+ dt.strftime('%y'))
    begin_month = int(begin_day.strftime('%m'))
    end_month = int(end_day.strftime('%m'))
    begin_day = int(begin_day.strftime('%d'))
    end_day = int(end_day.strftime('%d'))
    end_timelist = (current_year,end_month,end_day,0,0,0,0,0,0)
    begin_timelist = (current_year,begin_month,begin_day,0,0,0,0,0,0)
    endTime_str = str(int(time.mktime(end_timelist)))+'000'
    beginTime_str = str(int(time.mktime(begin_timelist)))+'000'
    timeBandE = [beginTime_str,endTime_str]
    return timeBandE

def getRemoteHistoryPostInfo(queueName,timeB,timeE):
    postDict2_getHistory = {
    'callCount':'1',
    'page':'/operation/ems/ems_history.jsf',
    'httpSessionId':'7cfb8c9094be950206736c4a6b128dd8d7c22f3440b1b7c22db97cbc9d88e05d.e38Pa30Pbx4PbO0PbhqPc3qKaN90',
    'scriptSessionId':'EDB6039197EB4A3136E0A0639F4CB372975',
    'c0-scriptName':'EmsHistoryRemote',
    'c0-methodName':'searchHistoryAjax',
    'c0-id':'0',
    'c0-param0':'string:%7B%22serverName%22%3A%22tcp%3A%2F%2Fcosems01v%3A8001%22%2C%22stype%22%3A%22queue%22%2C%22fromTime%22%3A'+timeB+'%2C%22toTime%22%3A'+timeE+'%2C%22queueName%22%3A%22'+ queueName +'%22%7D',
    'batchId':'0'
    }
    return postDict2_getHistory

def getJson(data_dict_post={}):
    global opener
    url_gethistory = 'https://www.cargosmart.com/operation/dwr/call/plaincall/EmsHistoryRemote.searchHistoryAjax.dwr'
    getHistory_post_send = urllib.parse.urlencode(data_dict_post).encode()
    op = opener.open(url_gethistory,getHistory_post_send)
    data = op.read().decode('utf-8')
    json_data_list = re.findall(r'inboundTotalMessages\\":(.+?),\\',data)
    return json_data_list

##################################################
    
url = 'https://www.cargosmart.com'

header = {
'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0',
'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
'Accept-Language':'en-US,en;q=0.5',
'Accept-Encoding':'gzip, deflate',
'Referer':"https://www.cargosmart.com/operation/login.html",
'Connection':"keep-alive",
'Host':"www.cargosmart.com"
}
ssl._create_default_https_context = ssl._create_unverified_context  #处理证书
opener = getOpener(header)

def main_crawler():
    global url
    global opener

    #Login info as a dict to post
    postDict = {
    'j_username':'ocuser',
    'j_password':'ocuser'
    }
    
    postData = urllib.parse.urlencode(postDict).encode()
    
    url_login = url + '/operation/j_spring_security_check'
    op = opener.open(url_login,postData)
    data = op.read().decode('utf-8')
    
    #用正则表达式筛选网址
    linkre = re.compile('/operation/ems/ems_history.jsf')
    for x in linkre.findall(data):
        url_history = url + x
    op = opener.open(url_history)
    data = op.read().decode('utf-8')

    queueName_BC = 'Cargosmart.COSCON.CD.BC.INPUT.PARALLEL'
    queueName_CT = 'Cargosmart.COSCON.CD.CT.INPUT.PARALLEL'
    queueName_BL = 'Cargosmart.COSCON.CD.BL.INPUT.PARALLEL'
    queueName_BLPrint = 'CS2.BILLOFLADING.PRINT.COSU.IN.QUE'

    days = input("How many days to retrospect?(1 is one day,2 are two days and so on):")
    timeCY = getTimeSec(days)
    timeB = timeCY[0]
    timeE = timeCY[1]

    #get BC
    BC_history_Post = getRemoteHistoryPostInfo(queueName_BC,timeB,timeE)
    History_list_BC = getJson(BC_history_Post)
    BC_history_count_Oneday = int(History_list_BC[len(History_list_BC)-1])-int(History_list_BC[0])

    #get BL
    BL_history_Post = getRemoteHistoryPostInfo(queueName_BL,timeB,timeE)
    History_list_BL = getJson(BL_history_Post)
    BL_history_count_Oneday = int(History_list_BL[len(History_list_BL)-1])-int(History_list_BL[0])

    #get CT
    CT_history_Post = getRemoteHistoryPostInfo(queueName_CT,timeB,timeE)
    History_list_CT = getJson(CT_history_Post)
    CT_history_count_Oneday = int(History_list_CT[len(History_list_CT)-1])-int(History_list_CT[0])

    #get BLPrint
    BLPrint_history_Post = getRemoteHistoryPostInfo(queueName_BLPrint,timeB,timeE)
    History_list_BLPrint = getJson(BLPrint_history_Post)
    BLPrint_history_count_Oneday = int(History_list_BLPrint[len(History_list_BLPrint)-1])-int(History_list_BLPrint[0])

    queuePrint = [CT_history_count_Oneday,BC_history_count_Oneday,BL_history_count_Oneday,BLPrint_history_count_Oneday]
    return queuePrint

def GetDaysRange():
    x = input('How many days to retrospect?(input 1 one day ago，input 2 two days ago，input 3 three days ago):')
    return x

def AutoGetColPosFator(daysRange):
    nowtime = datetime.datetime.now()
    needtime = nowtime + datetime.timedelta(days = 0-int(daysRange))
    #因为星期是0~6，所以为了通俗的显示，+1处理
    weekday = int(needtime.weekday())+1
    if (weekday+3) > 7:
        return weekday+3-7
    else:
        return weekday+3


def insert_UserSync_And_Shipments():
    tt = datetime.datetime.now()
    targetfilekeyword = tt.strftime('%m')+tt.strftime('%d')
    
    try:
        daysRange = GetDaysRange()
        fator = AutoGetColPosFator(daysRange)
        yt = tt + datetime.timedelta(days=0-int(daysRange))
        targetfolderkeyword = '20'+yt.strftime("%y")+yt.strftime("%m")+yt.strftime("%d")
        #打开Excel
        xls = UserSync_And_Shipments(r'D:\\DailyReportResouceFiles\\Report\\COSCON-ACZ-daily-stat-result '+targetfilekeyword+'.xlsx')

        if fator == 1:
        
            xls.changeCaptionDate('UserSync',tt,3,3,3,13,1)
            xls.changeCaptionDate('Shipments',tt,2,1,7,8,2)
            xls.changeCaptionDate('Shipments',tt,17,1,7,17,4)
           
        #填入user
        list_user = excel_get_user(r'D:\\DailyReportResouceFiles\\'+targetfolderkeyword+'\\Report - Coscon User Profile Sync Txn Report.xlsx')

        if fator ==1:
              xls.dealWithMondayUserTable('UserSync',list_user)
        if fator>1 and fator<4:
            xls.setCell(list_user,3,5,(7+(fator-1-1)*4),'UserSync')
        elif fator==4:
            xls.setCell(list_user,3,19,1,'UserSync')
        elif fator == 5:
            xls.setCell(list_user,3,19,7,'UserSync')
        elif fator == 6:
            xls.setCell(list_user,3,19,11,'UserSync')
        else:
            xls.setCell(list_user,3,32,1,'UserSync')
        
        #填入daily
        list_aczone = excel_get_zone_or_shipment(r'D:\\DailyReportResouceFiles\\'+targetfolderkeyword+'\\ACZone TXN Monitor.xlsx',0,u'Sheet0',daysRange)
        list_stdzone = excel_get_zone_or_shipment(r'D:\\DailyReportResouceFiles\\'+targetfolderkeyword+'\\STDZone COSCON BR SI Daily TXN Report.xlsx',0,u'Sheet0',daysRange)
        list_daily_result = calc(list_aczone,list_stdzone)[0]    
        xls.setCell(list_daily_result,1,19,(4+(fator-1)*2),'Shipments')

        #填入shipment
        list_shipment = excel_get_zone_or_shipment(r'D:\\DailyReportResouceFiles\\'+targetfolderkeyword+'\\ACZone Shipment Folder Txn Report.xlsx',0,u'Sheet0',daysRange)
        e_list_shipment = [0]
        if list_shipment:
            e_list_shipment[0] = list_shipment[1]
        else:
            e_list_shipment[0] = 0   
        xls.setCell(e_list_shipment,1,23,(4+(fator-1)*2),'Shipments')

        #填入计算后的值
        list_calc_result = calc(list_aczone,list_stdzone)[1]
        xls.setCell(list_calc_result,1,8,(2+(fator-1)),'Shipments')

##        #填入ct，bl，bc值
##        list_queue = main_crawler()
##        xls.setCell(list_queue,1,24,(4+(fator-1)*2),'Shipments')
##        xls.setCell(list_queue,1,3,(2+(fator-1)),'Shipments')
        
    except Exception(e):
        print(str(e))
        xls.close()
   
    xls.save()
    xls.close()






cur_time=time.time()
yes_time=cur_time-86400
yesterday= time.strftime("%Y%m%d",time.localtime(yes_time))
today=time.strftime("%m%d",time.localtime(time.time()))#不要修改time.time()
ytd=time.strftime("%a",time.localtime(yes_time))
pic1=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'COSCON Network Utilization','5min.png'])
pic2=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'COSCON Network Utilization','30min.png'])
excelPath=r'\\'.join(['D:','DailyReportResouceFiles','Report','COSCON-ACZ-daily-stat-result '+today+'.xlsx'])
#excelPath2="D:\DailyReportResouceFiles\Report\COSCON-ACZ-daily-stat-result "+today+".xlsx"
insert_date=time.strftime("%Y/%b/%d",time.localtime(yes_time))+' COSCON 10M lease line usage : < 25%'
yestd= time.strftime("%Y%m%d",time.localtime(yes_time))
rootdir = "D:/DailyReportResouceFiles/"+yestd# 指明压缩文件路径
zipdir = "D:/DailyReportResouceFiles/"+yestd+"/COSCON Network Utilization"    # 存储解压缩后的文件夹
      
def Mon(xls):
      xls.addPicture('Network', pic1, 0,500,389,170)
      xls.addPicture('Network', pic2, 0,735,389,170)
      xls.setCell(32,1,insert_date)
      xls.setCell(47,3,'Daily (5 minutes average)')
      xls.setCell(62,3,'Weekly (30 minutes average)')
      save_close(xls)
def Tue(xls):
      xls.addPicture('Network', pic1, 480,500,389,170)
      xls.addPicture('Network', pic2, 480,735,389,170)
      xls.setCell(32,11,insert_date)
      xls.setCell(47,13,'Daily (5 minutes average)')
      xls.setCell(62,13,'Weekly (30 minutes average)')
      save_close(xls)
def Wed(xls):
      xls.addPicture('Network', pic1, 960,500,389,170)
      xls.addPicture('Network', pic2, 960,735,389,170)
      xls.setCell(32,21,insert_date)
      xls.setCell(47,23,'Daily (5 minutes average)')
      xls.setCell(62,23,'Weekly (30 minutes average)')
      save_close(xls)
def Thu(xls):
      xls.addPicture('Network', pic1, 0,975,389,170)
      xls.addPicture('Network', pic2, 0,1210,389,170)
      xls.setCell(64,1,insert_date)
      xls.setCell(78,3,'Daily (5 minutes average)')
      xls.setCell(94,3,'Weekly (30 minutes average)')
      save_close(xls)
def Fri(xls):
      clearSheet(xls)
      xls = insert_Picture(excelPath)
      xls.addPicture('Network', pic1, 0,25,389,170)
      xls.addPicture('Network', pic2, 0,260,389,170)
      xls.setCell(1,1,insert_date)
      xls.setCell(15,3,'Daily (5 minutes average)')
      xls.setCell(30,3,'Weekly (30 minutes average)')
      save_close(xls)
def Sat(xls):
      xls.addPicture('Network', pic1, 480,25,389,170)
      xls.addPicture('Network', pic2, 480,260,389,170)
      xls.setCell(1,11,insert_date)
      xls.setCell(15,13,'Daily (5 minutes average)')
      xls.setCell(30,13,'Weekly (30 minutes average)')
      save_close(xls)
def Sun(xls):
      xls.addPicture('Network', pic1, 960,25,389,170)
      xls.addPicture('Network', pic2, 960,260,389,170)
      xls.setCell(1,21,insert_date)
      xls.setCell(15,23,'Daily (5 minutes average)')
      xls.setCell(30,23,'Weekly (30 minutes average)')
      yesterday= time.strftime("%Y%m%d",time.localtime(yes_time))
      BCPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'BC2.png'])
      BLPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'BL2.png'])
      CTPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'CT2.png'])
      xls.addPicture('Throughout', BCPath, 0,20,1230,270)
      xls.addPicture('Throughout', BLPath, 0,380,1230,270)
      xls.addPicture('Throughout', CTPath, 0,740,1230,270)
      save_close(xls)
def clearSheet(xls):
      clearRangeData(xls,'UserSync', 5, 7, 16, 13)
      clearRangeData(xls,'UserSync', 19, 1, 29, 13)
      clearRangeData(xls,'UserSync', 32, 1, 45, 3)
      #更新时间
      setDateCell(xls,'UserSync',3,1,time.strftime("%m/%d/%Y",time.localtime(time.time()-(86400*3))))
      setDateCell(xls,'UserSync',3,7,time.strftime("%m/%d/%Y",time.localtime(time.time()-(86400*2))))
      setDateCell(xls,'UserSync',3,11,time.strftime("%m/%d/%Y",time.localtime(time.time()-(86400*1))))
      setDateCell(xls,'UserSync',17,1,time.strftime("%m/%d/%Y",time.localtime(time.time()+(86400*0))))
      setDateCell(xls,'UserSync',17,7,time.strftime("%m/%d/%Y",time.localtime(time.time()+(86400*1))))
      setDateCell(xls,'UserSync',17,11,time.strftime("%m/%d/%Y",time.localtime(time.time()+(86400*2))))
      setDateCell(xls,'UserSync',30,1,time.strftime("%m/%d/%Y",time.localtime(time.time()+(86400*3))))
      clearRangeData(xls,'Shipments', 3, 2, 12, 8)
      clearRangeData(xls,'Shipments', 19, 4, 27, 17)
      save_close(xls)
      time.sleep(1)
      try:
            #wb = openpyxl.reader.excel.load_workbook(excelPath)
            wb = openpyxl.load_workbook(excelPath)
            wb.remove_sheet(wb.get_sheet_by_name('Throughout'))
            wb.remove_sheet(wb.get_sheet_by_name('Network'))#清空第Network
            wb.remove_sheet(wb.get_sheet_by_name('ServerPerformance'))
            wb.create_sheet("Throughout", 2)
            wb.create_sheet("ServerPerformance", 3)
            wb.create_sheet("Network", 4)
            wb.save(excelPath)
      except Exception(e):
            print(str(e))



def setDateCell(xls, sheet, row, col, value):  #设置单元格的数据  
    "set value of one cell"    
    sht = xls.xlBook.Worksheets(sheet)    
    sht.Cells(row, col).Value = value 


def save_close(xls):
     xls.save()
     xls.close()

def clearRangeData(xls, sheet, row1, col1, row2, col2):  #获得一块区域的数据，返回为一个二维元组,并清除为空  
    #"return a 2d array (i.e. tuple of tuples)"    
    sht = xls.xlBook.Worksheets(sheet)  
    sht.Range(sht.Cells(row1, col1), sht.Cells(row2, col2)).Value = ''

def write_performance():
    dt = datetime.datetime.now()
    ydt = dt + datetime.timedelta(days = -1)
    sourceFileKeyword = '20'+ydt.strftime("%y")+ydt.strftime("%m")+ydt.strftime("%d")
    targetFilekeyword = dt.strftime("%m")+dt.strftime("%d")
    
    opener = openCSV()
    
    ##获取COSCONACZ的数据
    csvData_cosconacz = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCONACZ-PROD.csv")
    csvDataList_cosconacz = opener.getlist(csvData_cosconacz)
    csvData_cosconacz = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCONACZ-PROD.csv")
    multiplier_cosconacz = opener.getMaxColNumber(csvData_cosconacz)
    ##获取COSCON的数据
    csvData_coscon = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCON-PROD.csv")
    csvDataList_coscon = opener.getlist(csvData_coscon)
    csvData_coscon = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCON-PROD.csv")
    multiplier_coscon = opener.getMaxColNumber(csvData_coscon)

    xls = InputExcel(r"D:\\DailyReportResouceFiles\\Report\\COSCON-ACZ-daily-stat-result "+targetFilekeyword+".xlsx")
    theLastRowPos = xls.getTheLastRowPos('ServerPerformance')
    xls.setDateCell(theLastRowPos,'ServerPerformance')
    
    if int(multiplier_cosconacz) ==  int(multiplier_coscon) or int(multiplier_cosconacz) > int(multiplier_coscon):
        acz_row = xls.setCell(csvDataList_cosconacz,multiplier_cosconacz,theLastRowPos+2,1,'ServerPerformance')
        coscon_row = xls.setCell(csvDataList_coscon[multiplier_coscon:],multiplier_coscon,theLastRowPos+2+acz_row,1,'ServerPerformance')

    else:
        coscon_row = xls.setCell(csvDataList_coscon,multiplier_coscon,theLastRowPos+2,1,'ServerPerformance')
        acz_row = xls.setCell(csvDataList_cosconacz[multiplier_cosconacz:],multiplier_cosconacz,theLastRowPos+2+coscon_row,1,'ServerPerformance')
    xls.save()
    xls.close()



#解压缩Zip到指定文件夹
def extractZip(zfile, path):
    z = ZFile(zfile)
    z.extract_to(path)
    z.close()

#获得文件名和后缀
def GetFileNameAndExt(filename):
    (filepath,tempfilename) = os.path.split(filename);
    (shotname,extension) = os.path.splitext(tempfilename);
    return shotname,extension

#定义文件处理数量-全局变量
fileCount = 0

#递归获得zip文件集合
def getFiles(filepath):
  #遍历filepath下所有文件，包括子目录
  files = os.listdir(filepath)
  for fi in files:
    fi_d = os.path.join(filepath,fi)
    if os.path.isdir(fi_d):
        getFiles(fi_d)
    else:
        global fileCount
        global zipdir
        fileCount = fileCount + 1
        # print fileCount
        fileName = os.path.join(filepath,fi_d)
        filenamenoext = GetFileNameAndExt(fileName)[0]
        fileext = GetFileNameAndExt(fileName)[1]
        # 如果要保存到同一个文件夹，将文件名设为空
        filenamenoext = ""
        zipdirdest = zipdir + "/" + filenamenoext + "/"
        if fileext in ['.zip']:
            if not os.path.isdir(zipdirdest):
                os.mkdir(zipdirdest)
        if fileext == ".zip" :#
            print (str(fileCount) + " -- " + fileName)
           # unzip(fileName,zipdirdest)
            extractZip(fileName,zipdirdest)



class openCSV:
    
    def openFile(self,FileName):
        try:
            self.csv_reader = csv.reader(open(FileName,encoding='utf-8'))
        except Exception(e):
            print(str(e))
        return self.csv_reader

    def getlist(self,csvData):
        self.csvDataList = []
        for row in csvData:
            for col in range(len(row)):
                self.csvDataList.append(row[col])
        return self.csvDataList
    
    def getMaxColNumber(self,csvData):
        self.TableCaption = []
        for row in csvData:
            for col in range(len(row)):
                self.TableCaption.append(row[col])
            break
        return len(self.TableCaption)

class InputExcel:
    def __init__(self,filename=None):
        self.xlApp = win32com.client.Dispatch('Excel.Application')
        if filename:
            self.filename = filename
            self.xlBook = self.xlApp.Workbooks.Open(filename)
        else:
            self.xlBook = self.xlApp.Workbooks.Add()
            self.filename = ''

    '''
    lists:用数组作数据源
    multiplier:乘数，就是原数据每一行的列数
    row_in_sheet:在导入的表里开始位置的行坐标
    col_in_sheet:在导入的表里开始位置的列坐标
    sheet:创的表格
    '''
    def setCell(self,lists,multiplier,row_in_sheet,col_in_sheet,sheet):
        sht = self.xlBook.Worksheets(sheet)
        begin = 0
        end = 0
        ListLength = int(len(lists))
        if ListLength == 0:
              rows = 0
        else:
              rows = ListLength/multiplier
              for i in range(int(rows)):
                  end = end + multiplier
                  e_list = lists[begin:end]
                  for j in range(len(e_list)):
                      #sheet.write(row_in_sheet,j+col_in_sheet,e_list[j])
                      sht.Cells(row_in_sheet,j+col_in_sheet).BorderAround(1,2,3)
                      sht.Columns(j+col_in_sheet).ColumnWidth=40
                      sht.Rows(row_in_sheet).RowHeight = 28
                      d = sht.Cells(row_in_sheet, j+col_in_sheet)
                      d.Value = (e_list[j])
                  row_in_sheet = row_in_sheet + 1
                  begin = begin + multiplier
        return rows

    def setDateCell(self,theLastRowPos,sheet):
        self.sht = self.xlBook.Worksheets(sheet)
        self.dt = datetime.datetime.now()
        self.Ydt = self.dt + datetime.timedelta(days = -1)
        self.Ytyear = '20'+self.Ydt.strftime("%y")
        self.Ytmon = self.Ydt.strftime("%m")
        self.Ytday = self.Ydt.strftime("%d")
        self.timeStr = self.Ytyear+'-'+self.Ytmon+'-'+self.Ytday
        self.sht.Rows(theLastRowPos).RowHeight = 28
        self.sht.Rows(theLastRowPos+1).RowHeight = 28
        self.d = self.sht.Cells(theLastRowPos+1,1)
        self.d.Value = self.timeStr+' 00:00:00  to '+self.timeStr+' 23:59:59 HKT'

    def getTheLastRowPos(self,sheet):
        self.sht = self.xlBook.Worksheets(sheet)
        self.number = 1
        while 1:
            self.cell = self.sht.Cells(self.number,1).value
            self.cellNext = self.sht.Cells(self.number+2,1).value
            if self.cell or self.cellNext:
                self.number = self.number + 1
            else:
                break
        return self.number

    def save(self,newfilename=None):
        if newfilename:
            self.filename = newfilename
            self.xlBook.SaveAs(newfilename)
        else:
            self.xlBook.Save()
    def close(self):
        self.xlBook.Close(SaveChanges=0)
        del self.xlApp

class insert_Picture:  
      def __init__(self, filename=None):  #打开文件或者新建文件（如果不存在的话）
          self.xlApp = win32com.client.Dispatch('Excel.Application')  
          if filename:  
              self.filename = filename  
              self.xlBook = self.xlApp.Workbooks.Open(filename)  
          else:  
              self.xlBook = self.xlApp.Workbooks.Add()  
              self.filename = ''
      def save(self, newfilename=None):  #保存文件
          if newfilename:  
              self.filename = newfilename  
              self.xlBook.SaveAs(newfilename)  
          else:  
              self.xlBook.Save()      
      def close(self):  #关闭文件
          self.xlBook.Close(SaveChanges=0)  
          del self.xlApp  
      def addPicture(self, sheet, pictureName, Left, Top, Width, Height):  #插入图片
          sht = self.xlBook.Worksheets(sheet)  
          sht.Shapes.AddPicture(pictureName, 1, 1, Left, Top, Width, Height)
      def setCell(self,row,col,value):  #设置单元格的数据
          sht = self.xlBook.Worksheets('Network')
          sht.Cells(row, col).Value = value



#Zip文件处理类
class ZFile(object):
    def __init__(self, filename, mode='r', basedir=''):
        self.filename = filename
        self.mode = mode
        if self.mode in ('w', 'a'):
            self.zfile = zipfile.ZipFile(filename, self.mode, compression=zipfile.ZIP_DEFLATED)
        else:
            self.zfile = zipfile.ZipFile(filename, self.mode)
        self.basedir = basedir
        if not self.basedir:
            self.basedir = os.path.dirname(filename)

    def addfile(self, path, arcname=None):
        path = path.replace('//', '/')
        if not arcname:
            if path.startswith(self.basedir):
                arcname = path[len(self.basedir):]
            else:
                arcname = ''
        self.zfile.write(path, arcname)

    def addfiles(self, paths):
        for path in paths:
            if isinstance(path, tuple):
                self.addfile(*path)
            else:
                self.addfile(path)

    def close(self):
        self.zfile.close()

    def extract_to(self, path):
        for p in self.zfile.namelist():
            self.extract(p, path)

    def extract(self, filename, path):
        if not filename.endswith('/'):
            f = os.path.join(path, filename)
            dir = os.path.dirname(f)
            if not os.path.exists(dir):
                os.makedirs(dir)
            open(f, 'wb').write(self.zfile.read(filename))



class UserSync_And_Shipments:
    def __init__(self,filename=None):
        self.xlApp = win32com.client.Dispatch('Excel.Application')
        if filename:
            self.filename = filename
            self.xlBook = self.xlApp.Workbooks.Open(filename)
        else:
            self.xlBook = self.xlApp.Workbooks.Add()
            self.filename = ''

    def save(self,newfilename=None):
        if newfilename:
            self.filename = newfilename
            self.xlBook.SaveAs(newfilename)
        else:
            self.xlBook.Save()
    def close(self):
        self.xlBook.Close(SaveChanges=0)
        del self.xlApp

    '''
    lists:用数组作数据源
    multiplier:乘数，就是原数据每一行的列数
    row_in_sheet:在导入的表里开始位置的行坐标
    col_in_sheet:在导入的表里开始位置的列坐标
    sheet:创的表格
    '''
    def setCell(self,lists,multiplier,row_in_sheet,col_in_sheet,sheet):
        sht = self.xlBook.Worksheets(sheet)
        begin = 0
        end = 0
        rows = len(lists)/multiplier
        for i in range(int(rows)):
            end = end + multiplier
            e_list = lists[begin:end]
            for j in range(len(e_list)):
                #sheet.write(row_in_sheet,j+col_in_sheet,e_list[j])
                d = sht.Cells(row_in_sheet, j+col_in_sheet)
                d.Value = (e_list[j])
            row_in_sheet = row_in_sheet + 1
            begin = begin + multiplier

##@sheet:表名
##@nowdatetime:当前日期
##@beginRowPos:开始行的行号
##@MaxloopTimes:最大循环次数，比如user表需要插三行的数据，MaxloopTimes为3
##@ElemNumerARow:每一行要插入的元素个数
##@Maxlen:一行的最大长度
##@BeginColPos:开始列的位置
    def changeCaptionDate(self,sheet,nowdatetime,beginRowPos,MaxloopTimes,ElemNumberARow,Maxlen,BeginColPos):
        begindate = nowdatetime + datetime.timedelta(days = -3)       
        sht = self.xlBook.Worksheets(sheet)
        
        self.currentnumber = 1
        self.col_datepos = BeginColPos
        self.loopTimes = 1        
        while self.currentnumber <= ElemNumberARow:
        
            while self.col_datepos <= Maxlen: 
                d = sht.Cells(beginRowPos,self.col_datepos).Value
                if d:
                    self.targetdate = begindate + datetime.timedelta(days = (self.currentnumber-1)+3*(self.loopTimes-1))
                    sht.Cells(beginRowPos,self.col_datepos).Value = (self.targetdate.strftime("%m/%d/%y"))                  
                    self.col_datepos = self.col_datepos + 1

                    if self.currentnumber == ElemNumberARow and self.loopTimes < MaxloopTimes:
                        self.currentnumber = 0
                        self.loopTimes = self.loopTimes +1
                        beginRowPos = beginRowPos + 14
                        self.col_datepos = 1
                        break
                    else:
                        break
                else:
                    self.col_datepos = self.col_datepos + 1          
            self.currentnumber = self.currentnumber + 1
            

    def dealWithMondayUserTable(self,sheet,userlist):
        sht = self.xlBook.Worksheets(sheet)
        self.currentNeedFillNumber = 1
        self.currentCursorRow = 5
        self.currentCursorColumn = 1
  
        while self.currentNeedFillNumber <= 10:
            self.userList = userlist
            self.operationCode = sht.Cells(self.currentCursorRow,self.currentCursorColumn).Value
            self.returnResult = sht.Cells(self.currentCursorRow,self.currentCursorColumn+1).Value
            
            for Pointer_userList in range(len(self.userList)):
                
                if self.operationCode == self.userList[Pointer_userList] and self.returnResult == self.userList[Pointer_userList+1]:
                    sht.Cells(self.currentCursorRow,self.currentCursorColumn+2).Value = self.userList[Pointer_userList+2]               
                    break                
                elif Pointer_userList == (len(self.userList)-1):
                    sht.Cells(self.currentCursorRow,self.currentCursorColumn+2).Value = 0
            self.currentCursorRow = self.currentCursorRow + 1
            
            self.currentNeedFillNumber = self.currentNeedFillNumber+1




if __name__ == "__main__":
      getFiles(rootdir)
      time.sleep(1)
      if not os.path.isfile(pic1) or not os.path.isfile(excelPath):
            print('network图片或COSCON-ACZ-daily-stat-result.xlsx不存在')
            sys.exit()
      xls = insert_Picture(excelPath)
      result={"Mon":lambda:Mon(xls),
              "Tue":lambda:Tue(xls),
              "Wed":lambda:Wed(xls),
              "Thu":lambda:Thu(xls),
              "Fri":lambda:Fri(xls),
              "Sat":lambda:Sat(xls),
              "Sun":lambda:Sun(xls)
              }
      #result["Sun"]()
      result[ytd]()
      write_performance()
      insert_UserSync_And_Shipments()
      sys.exit()
     
