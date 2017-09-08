#!/usr/bin/env python   
# -*- coding: utf-8 -*- 
from win32com.client import Dispatch
import win32com.client
import time
import os
import os.path
import openpyxl
import sys
import csv
import datetime
import zipfile
from zipfile import *

yesterday= time.strftime("%Y%m%d",time.localtime(time.time()-86400))
today=time.strftime("%m%d",time.localtime(time.time()))
ytd=time.strftime("%a",time.localtime(time.time()-86400))
pic1=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'COSCON Network Utilization','5min.png'])
pic2=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'COSCON Network Utilization','30min.png'])
excelPath=r'\\'.join(['D:','DailyReportResouceFiles','Report','COSCON-ACZ-daily-stat-result '+today+'.xlsx'])
#excelPath2="D:\DailyReportResouceFiles\Report\COSCON-ACZ-daily-stat-result "+today+".xlsx"
insert_date=time.strftime("%Y/%b/%d",time.localtime(time.time()-86400))+' COSCON 10M lease line usage : < 25%'
yestd= time.strftime("%Y%m%d",time.localtime(time.time()-86400))
rootdir = "D:/DailyReportResouceFiles/"+yestd# 指明压缩文件路径
zipdir = "D:/DailyReportResouceFiles/"+yestd+"/COSCON Network Utilization"    # 存储解压缩后的文件夹
      
def Mon(xls):
      xls.addPicture('Network', pic1, 0,500,389,170)
      xls.addPicture('Network', pic2, 0,735,389,170)
      xls.setCell(32,1,insert_date)
      xls.setCell(47,3,'Daily (5 minutes average)')
      xls.setCell(62,3,'Weekly (30 minutes average)')
      save_close(xls)
def Tue(xls):
      xls.addPicture('Network', pic1, 480,500,389,170)
      xls.addPicture('Network', pic2, 480,735,389,170)
      xls.setCell(32,11,insert_date)
      xls.setCell(47,13,'Daily (5 minutes average)')
      xls.setCell(62,13,'Weekly (30 minutes average)')
      save_close(xls)
def Wed(xls):
      xls.addPicture('Network', pic1, 960,500,389,170)
      xls.addPicture('Network', pic2, 960,735,389,170)
      xls.setCell(32,21,insert_date)
      xls.setCell(47,23,'Daily (5 minutes average)')
      xls.setCell(62,23,'Weekly (30 minutes average)')
      save_close(xls)
def Thu(xls):
      xls.addPicture('Network', pic1, 0,975,389,170)
      xls.addPicture('Network', pic2, 0,1210,389,170)
      xls.setCell(64,1,insert_date)
      xls.setCell(78,3,'Daily (5 minutes average)')
      xls.setCell(94,3,'Weekly (30 minutes average)')
      save_close(xls)
def Fri(xls):
      clearSheet(xls)
      xls = insert_Picture(excelPath)
      xls.addPicture('Network', pic1, 0,25,389,170)
      xls.addPicture('Network', pic2, 0,260,389,170)
      xls.setCell(1,1,insert_date)
      xls.setCell(15,3,'Daily (5 minutes average)')
      xls.setCell(30,3,'Weekly (30 minutes average)')
      save_close(xls)
def Sat(xls):
      xls.addPicture('Network', pic1, 480,25,389,170)
      xls.addPicture('Network', pic2, 480,260,389,170)
      xls.setCell(1,11,insert_date)
      xls.setCell(15,13,'Daily (5 minutes average)')
      xls.setCell(30,13,'Weekly (30 minutes average)')
      save_close(xls)
def Sun(xls):
      xls.addPicture('Network', pic1, 960,25,389,170)
      xls.addPicture('Network', pic2, 960,260,389,170)
      xls.setCell(1,21,insert_date)
      xls.setCell(15,23,'Daily (5 minutes average)')
      xls.setCell(30,23,'Weekly (30 minutes average)')
      yesterday= time.strftime("%Y%m%d",time.localtime(time.time()-86400))
      BCPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'BC2.png'])
      BLPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'BL2.png'])
      CTPath=r'\\'.join(['D:','DailyReportResouceFiles',yesterday,'CT2.png'])
      xls.addPicture('Throughout', BCPath, 0,20,1230,270)
      xls.addPicture('Throughout', BLPath, 0,380,1230,270)
      xls.addPicture('Throughout', CTPath, 0,740,1230,270)
      save_close(xls)
def clearSheet(xls):
      save_close(xls)
      time.sleep(1)
      try:
            #wb = openpyxl.reader.excel.load_workbook(excelPath)
            wb = openpyxl.load_workbook(excelPath)
            wb.remove_sheet(wb.get_sheet_by_name('Throughout'))
            wb.remove_sheet(wb.get_sheet_by_name('Network'))#清空第Network
            wb.remove_sheet(wb.get_sheet_by_name('ServerPerformance'))
            wb.create_sheet("Throughout", 2)
            wb.create_sheet("ServerPerformance", 3)
            wb.create_sheet("Network", 4)
            wb.save(excelPath)
      except Exception(e):
            print(str(e))
        
def save_close(xls):
     xls.save()
     xls.close()

def write_performance():
    dt = datetime.datetime.now()
    ydt = dt + datetime.timedelta(days = -1)
    sourceFileKeyword = '20'+ydt.strftime("%y")+ydt.strftime("%m")+ydt.strftime("%d")
    targetFilekeyword = dt.strftime("%m")+dt.strftime("%d")
    
    opener = openCSV()
    
    ##获取COSCONACZ的数据
    csvData_cosconacz = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCONACZ-PROD.csv")
    csvDataList_cosconacz = opener.getlist(csvData_cosconacz)
    csvData_cosconacz = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCONACZ-PROD.csv")
    multiplier_cosconacz = opener.getMaxColNumber(csvData_cosconacz)
    ##获取COSCON的数据
    csvData_coscon = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCON-PROD.csv")
    csvDataList_coscon = opener.getlist(csvData_coscon)
    csvData_coscon = opener.openFile("D:\\DailyReportResouceFiles\\"+sourceFileKeyword+"\\CS2-ACZ-COSCON-PROD.csv")
    multiplier_coscon = opener.getMaxColNumber(csvData_coscon)

    xls = InputExcel(r"D:\\DailyReportResouceFiles\\Report\\COSCON-ACZ-daily-stat-result "+targetFilekeyword+".xlsx")
    theLastRowPos = xls.getTheLastRowPos('ServerPerformance')
    xls.setDateCell(theLastRowPos,'ServerPerformance')
    
    if int(multiplier_cosconacz) ==  int(multiplier_coscon) or int(multiplier_cosconacz) > int(multiplier_coscon):
        acz_row = xls.setCell(csvDataList_cosconacz,multiplier_cosconacz,theLastRowPos+2,1,'ServerPerformance')
        coscon_row = xls.setCell(csvDataList_coscon[multiplier_coscon:],multiplier_coscon,theLastRowPos+2+acz_row,1,'ServerPerformance')

    else:
        coscon_row = xls.setCell(csvDataList_coscon,multiplier_coscon,theLastRowPos+2,1,'ServerPerformance')
        acz_row = xls.setCell(csvDataList_cosconacz[multiplier_cosconacz:],multiplier_cosconacz,theLastRowPos+2+coscon_row,1,'ServerPerformance')
    xls.save()
    xls.close()



#解压缩Zip到指定文件夹
def extractZip(zfile, path):
    z = ZFile(zfile)
    z.extract_to(path)
    z.close()

#获得文件名和后缀
def GetFileNameAndExt(filename):
    (filepath,tempfilename) = os.path.split(filename);
    (shotname,extension) = os.path.splitext(tempfilename);
    return shotname,extension

#定义文件处理数量-全局变量
fileCount = 0

#递归获得zip文件集合
def getFiles(filepath):
  #遍历filepath下所有文件，包括子目录
  files = os.listdir(filepath)
  for fi in files:
    fi_d = os.path.join(filepath,fi)
    if os.path.isdir(fi_d):
        getFiles(fi_d)
    else:
        global fileCount
        global zipdir
        fileCount = fileCount + 1
        # print fileCount
        fileName = os.path.join(filepath,fi_d)
        filenamenoext = GetFileNameAndExt(fileName)[0]
        fileext = GetFileNameAndExt(fileName)[1]
        # 如果要保存到同一个文件夹，将文件名设为空
        filenamenoext = ""
        zipdirdest = zipdir + "/" + filenamenoext + "/"
        if fileext in ['.zip']:
            if not os.path.isdir(zipdirdest):
                os.mkdir(zipdirdest)
        if fileext == ".zip" :#
            print (str(fileCount) + " -- " + fileName)
           # unzip(fileName,zipdirdest)
            extractZip(fileName,zipdirdest)



class openCSV:
    
    def openFile(self,FileName):
        try:
            self.csv_reader = csv.reader(open(FileName,encoding='utf-8'))
        except Exception(e):
            print(str(e))
        return self.csv_reader

    def getlist(self,csvData):
        self.csvDataList = []
        for row in csvData:
            for col in range(len(row)):
                self.csvDataList.append(row[col])
        return self.csvDataList
    
    def getMaxColNumber(self,csvData):
        self.TableCaption = []
        for row in csvData:
            for col in range(len(row)):
                self.TableCaption.append(row[col])
            break
        return len(self.TableCaption)

class InputExcel:
    def __init__(self,filename=None):
        self.xlApp = win32com.client.Dispatch('Excel.Application')
        if filename:
            self.filename = filename
            self.xlBook = self.xlApp.Workbooks.Open(filename)
        else:
            self.xlBook = self.xlApp.Workbooks.Add()
            self.filename = ''

    '''
    lists:用数组作数据源
    multiplier:乘数，就是原数据每一行的列数
    row_in_sheet:在导入的表里开始位置的行坐标
    col_in_sheet:在导入的表里开始位置的列坐标
    sheet:创的表格
    '''
    def setCell(self,lists,multiplier,row_in_sheet,col_in_sheet,sheet):
        sht = self.xlBook.Worksheets(sheet)
        begin = 0
        end = 0
        rows = len(lists)/multiplier
        for i in range(int(rows)):
            end = end + multiplier
            e_list = lists[begin:end]
            for j in range(len(e_list)):
                #sheet.write(row_in_sheet,j+col_in_sheet,e_list[j])
                sht.Cells(row_in_sheet,j+col_in_sheet).BorderAround(1,2,3)
                sht.Columns(j+col_in_sheet).ColumnWidth=40
                sht.Rows(row_in_sheet).RowHeight = 28
                d = sht.Cells(row_in_sheet, j+col_in_sheet)
                d.Value = (e_list[j])
            row_in_sheet = row_in_sheet + 1
            begin = begin + multiplier
        return rows

    def setDateCell(self,theLastRowPos,sheet):
        self.sht = self.xlBook.Worksheets(sheet)
        self.dt = datetime.datetime.now()
        self.Ydt = self.dt + datetime.timedelta(days = -1)
        self.Ytyear = '20'+self.Ydt.strftime("%y")
        self.Ytmon = self.Ydt.strftime("%m")
        self.Ytday = self.Ydt.strftime("%d")
        self.timeStr = self.Ytyear+'-'+self.Ytmon+'-'+self.Ytday
        self.sht.Rows(theLastRowPos).RowHeight = 28
        self.sht.Rows(theLastRowPos+1).RowHeight = 28
        self.d = self.sht.Cells(theLastRowPos+1,1)
        self.d.Value = self.timeStr+' 00:00:00  to '+self.timeStr+' 23:59:59 HKT'

    def getTheLastRowPos(self,sheet):
        self.sht = self.xlBook.Worksheets(sheet)
        self.number = 1
        while 1:
            self.cell = self.sht.Cells(self.number,1).value
            self.cellNext = self.sht.Cells(self.number+2,1).value
            if self.cell or self.cellNext:
                self.number = self.number + 1
            else:
                break
        return self.number

    def save(self,newfilename=None):
        if newfilename:
            self.filename = newfilename
            self.xlBook.SaveAs(newfilename)
        else:
            self.xlBook.Save()
    def close(self):
        self.xlBook.Close(SaveChanges=0)
        del self.xlApp

class insert_Picture:  
      def __init__(self, filename=None):  #打开文件或者新建文件（如果不存在的话）
          self.xlApp = win32com.client.Dispatch('Excel.Application')  
          if filename:  
              self.filename = filename  
              self.xlBook = self.xlApp.Workbooks.Open(filename)  
          else:  
              self.xlBook = self.xlApp.Workbooks.Add()  
              self.filename = ''
      def save(self, newfilename=None):  #保存文件
          if newfilename:  
              self.filename = newfilename  
              self.xlBook.SaveAs(newfilename)  
          else:  
              self.xlBook.Save()      
      def close(self):  #关闭文件
          self.xlBook.Close(SaveChanges=0)  
          del self.xlApp  
      def addPicture(self, sheet, pictureName, Left, Top, Width, Height):  #插入图片
          sht = self.xlBook.Worksheets(sheet)  
          sht.Shapes.AddPicture(pictureName, 1, 1, Left, Top, Width, Height)
      def setCell(self,row,col,value):  #设置单元格的数据
          sht = self.xlBook.Worksheets('Network')
          sht.Cells(row, col).Value = value



#Zip文件处理类
class ZFile(object):
    def __init__(self, filename, mode='r', basedir=''):
        self.filename = filename
        self.mode = mode
        if self.mode in ('w', 'a'):
            self.zfile = zipfile.ZipFile(filename, self.mode, compression=zipfile.ZIP_DEFLATED)
        else:
            self.zfile = zipfile.ZipFile(filename, self.mode)
        self.basedir = basedir
        if not self.basedir:
            self.basedir = os.path.dirname(filename)

    def addfile(self, path, arcname=None):
        path = path.replace('//', '/')
        if not arcname:
            if path.startswith(self.basedir):
                arcname = path[len(self.basedir):]
            else:
                arcname = ''
        self.zfile.write(path, arcname)

    def addfiles(self, paths):
        for path in paths:
            if isinstance(path, tuple):
                self.addfile(*path)
            else:
                self.addfile(path)

    def close(self):
        self.zfile.close()

    def extract_to(self, path):
        for p in self.zfile.namelist():
            self.extract(p, path)

    def extract(self, filename, path):
        if not filename.endswith('/'):
            f = os.path.join(path, filename)
            dir = os.path.dirname(f)
            if not os.path.exists(dir):
                os.makedirs(dir)
            open(f, 'wb').write(self.zfile.read(filename))


if __name__ == "__main__":
      getFiles(rootdir)
      time.sleep(1)
      if not os.path.isfile(pic1) or not os.path.isfile(excelPath):
            print('network图片或COSCON-ACZ-daily-stat-result.xlsx不存在')
            sys.exit()
      xls = insert_Picture(excelPath)
      result={"Mon":lambda:Mon(xls),
              "Tue":lambda:Tue(xls),
              "Wed":lambda:Wed(xls),
              "Thu":lambda:Thu(xls),
              "Fri":lambda:Fri(xls),
              "Sat":lambda:Sat(xls),
              "Sun":lambda:Sun(xls)
              }
      result["Fri"]()
      #result[ytd]()
      write_performance()
      sys.exit()
     
